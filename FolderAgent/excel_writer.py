"""
Excel writer for creating and updating Excel sheets from LLM JSON responses.
"""
import os
from openpyxl import Workbook, load_workbook
from FolderAgent.config import OUTPUT

class ExcelWriter:
    def __init__(self, output_file=None):
        self.output_file = output_file or OUTPUT
        
    def process_subfolder(self, subfolder_name, json_response):
        try:
            print(f"Processing subfolder: {subfolder_name}")
            print(f"JSON response: {json_response}")
            """Process subfolder - create Excel if first time, add row if exists."""
            if not os.path.exists(self.output_file):
                print(f"Creating new Excel file: {self.output_file}")
                # First time: create Excel with headers and first row
                self._create_excel_with_data(subfolder_name, json_response)
            else:
                print(f"Adding row to existing Excel: {self.output_file}")
                # Add new row to existing Excel
                self._add_row(subfolder_name, json_response)
        except Exception as e:
            print(f"Error processing subfolder {subfolder_name}: {e}")
            raise
    
    def _create_excel_with_data(self, subfolder_name, json_response):
        """Create new Excel with headers and first data row."""
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Folder Analysis"
            
            # Headers: Subfolder + JSON keys
            headers = ["Subfolder"] + list(json_response.keys())
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            # First data row
            worksheet.cell(row=2, column=1, value=subfolder_name)
            for col, value in enumerate(json_response.values(), 2):
                worksheet.cell(row=2, column=col, value=value)
            
            workbook.save(self.output_file)
            workbook.close()
        except Exception as e:
            print(f"Error creating Excel file: {e}")
            raise
    
    def _add_row(self, subfolder_name, json_response):
        """Add new row to existing Excel."""
        try:
            workbook = load_workbook(self.output_file)
            worksheet = workbook.active
            
            # Find next empty row
            next_row = worksheet.max_row + 1
            
            # Add subfolder name and JSON values
            worksheet.cell(row=next_row, column=1, value=subfolder_name)
            for col, value in enumerate(json_response.values(), 2):
                worksheet.cell(row=next_row, column=col, value=value)
            
            workbook.save(self.output_file)
            workbook.close()
        except Exception as e:
            print(f"Error adding row to Excel: {e}")
            raise
